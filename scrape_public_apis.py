import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from fpdf import FPDF

# ----------------------------------------
# CONFIGURATION
# ----------------------------------------

GITHUB_URL = "https://github.com/public-apis/public-apis"
EXCEL_FILE = "public_apis.xlsx"
PDF_FILE = "public_apis.pdf"

# ----------------------------------------
# UTILITY: Sanitize text for PDF encoding
# ----------------------------------------

def safe_text(text: str) -> str:
    """
    Converts problematic Unicode characters to ASCII-safe alternatives,
    compatible with FPDF's latin-1 encoding.
    """
    return (
        text.replace("’", "'")
            .replace("“", '"')
            .replace("”", '"')
            .replace("–", "-")
            .replace("—", "-")
            .replace("…", "...")
            .encode("latin-1", errors="replace")
            .decode("latin-1")
    )

# ----------------------------------------
# STEP 1: SCRAPE PUBLIC APIs FROM GITHUB
# ----------------------------------------

def scrape_public_apis():
    """
    Scrapes API category tables from the GitHub README page.
    Returns:
        headers: list of column names
        data_rows: list of API rows (each a list of column values)
    """
    response = requests.get(GITHUB_URL)
    soup = BeautifulSoup(response.text, "html.parser")
    tables = soup.select("article table")

    data_rows = []
    for table in tables:
        rows = table.select("tr")
        headers = [th.get_text(strip=True) for th in rows[0].select("th")]
        for row in rows[1:]:
            cols = [td.get_text(strip=True) for td in row.select("td")]
            if len(cols) == len(headers):
                data_rows.append(cols)

    return headers, data_rows

# ----------------------------------------
# STEP 2: EXPORT TO EXCEL (.xlsx)
# ----------------------------------------

def export_to_excel(headers, data, filename):
    """
    Writes the scraped API data to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Public APIs"

    # Write headers and data
    ws.append(headers)
    for row in data:
        ws.append(row)

    wb.save(filename)
    print(f"✅ Excel file saved as: {filename}")

# ----------------------------------------
# STEP 3: EXPORT TO PDF (Preview Table)
# ----------------------------------------

class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "Public APIs List (Preview)", ln=True, align="C")
        self.ln(5)

    def add_table(self, headers, data):
        self.set_font("Arial", "B", 10)
        for header in headers:
            self.cell(40, 8, safe_text(header)[:35], border=1)
        self.ln()

        self.set_font("Arial", "", 9)
        for row in data[:100]:  # Preview only first 100 APIs
            for col in row:
                self.cell(40, 8, safe_text(str(col))[:35], border=1)
            self.ln()

def export_to_pdf(headers, data, filename):
    """
    Creates a preview PDF of the scraped API data.
    Limits to 100 rows for formatting reasons.
    """
    pdf = PDF()
    pdf.add_page()
    pdf.add_table(headers, data)
    pdf.output(filename)
    print(f"✅ PDF file saved as: {filename}")

# ----------------------------------------
# MAIN EXECUTION
# ----------------------------------------

if __name__ == "__main__":
    print("🔎 Scraping public APIs from GitHub...")
    headers, data_rows = scrape_public_apis()

    print("📤 Exporting to Excel and PDF...")
    export_to_excel(headers, data_rows, EXCEL_FILE)
    export_to_pdf(headers, data_rows, PDF_FILE)

    print("\n🎉 Done! Both files are ready.")
